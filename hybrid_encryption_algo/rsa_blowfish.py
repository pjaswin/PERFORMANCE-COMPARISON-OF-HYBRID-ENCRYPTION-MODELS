from Crypto.Cipher import Blowfish, PKCS1_OAEP
from Crypto.PublicKey import RSA
import hashlib, secrets, binascii
import time
import psutil
from memory_profiler import profile
from collections import Counter
import math,sys
import openpyxl
import tracemalloc



# @profile
def encrypt_Blowfish_CBC(msg, secretKey):
    iv = secrets.token_bytes(8)
    cipher = Blowfish.new(secretKey, Blowfish.MODE_CBC, iv)
    padding_length = 8 - (len(msg) % 8)
    padded_msg = msg + bytes([padding_length] * padding_length)
    ciphertext = cipher.encrypt(padded_msg)
    return (ciphertext, iv)

# @profile
def decrypt_Blowfish_CBC(ciphertext, iv, secretKey):
    cipher = Blowfish.new(secretKey, Blowfish.MODE_CBC, iv)
    padded_msg = cipher.decrypt(ciphertext)
    padding_length = padded_msg[-1]
    return padded_msg[:-padding_length]


# def rsa_key_to_256_bit_key(key):
#     sha = hashlib.sha256(key.export_key())
#     return sha.digest()
# @profile
def encrypt_RSA(msg, pubKey):
    rsaCipher = PKCS1_OAEP.new(pubKey)
    secretKey = secrets.token_bytes(32)
    ciphertext,iv = encrypt_Blowfish_CBC(msg, secretKey)
    ciphertextKey = rsaCipher.encrypt(secretKey)
    return (ciphertext,iv, ciphertextKey)

# @profile
def decrypt_RSA(encryptedMsg, privKey):
    (ciphertext, iv,ciphertextKey) = encryptedMsg
    rsaCipher = PKCS1_OAEP.new(privKey)
    secretKey = rsaCipher.decrypt(ciphertextKey)
    plaintext = decrypt_Blowfish_CBC(ciphertext, iv,secretKey)
    return plaintext

count=0
workbook = openpyxl.Workbook()
sheet = workbook.active
outputs=["file name","encry_time","decry_time","encry_cpu","decry_cpu","encry_mem","decry_mem","Hamming Distance","Avalache Effect","Through_put encryt","Through_put decrypt"]
file_list=["oneMB.txt","twoMB.txt","fiveMB.txt","tenMB.txt","twentyMB.txt"]
file_count=0
row_count=2
for column, output in enumerate(outputs, start=1):
    sheet.cell(row=1, column=column, value=output)
while file_count<5:
    count=0
    while count<30:

        with open(file_list[file_count], 'r') as file:
            msg = file.read()
        process = psutil.Process()
        modified_msg = msg[:2] + 'XY' + msg[4:]
        msg=msg.encode()
        modified_msg=modified_msg.encode()
        # msg = b'Text to be encrypted by ECC public key and ' \
        #       b'decrypted by its corresponding ECC private key'
        # print("original msg:", msg)
        tracemalloc.start()
        current_encrpt, peak_encrpt = tracemalloc.get_traced_memory()
        start_cpu=process.cpu_percent()
        # start_cpu = psutil.cpu_percent()
        start_time = time.perf_counter()

        key = RSA.generate(2048)
        privKey = key
        pubKey = key.publickey()
    
        encryptedMsg = encrypt_RSA(msg, pubKey)
        end_time = time.perf_counter()
        encry_time=end_time-start_time
        end_cpu=process.cpu_percent()
        # end_cpu = psutil.cpu_percent()
        encry_cpu=end_cpu-start_cpu
        current_encrpt2, peak_encrpt2 = tracemalloc.get_traced_memory()
        mem_usage=current_encrpt2-current_encrpt
        encryptedMsgObj = {
            'ciphertext': binascii.hexlify(encryptedMsg[0]),
            'iv': binascii.hexlify(encryptedMsg[1]),
            'ciphertextKey': binascii.hexlify(encryptedMsg[2])
        }
        # print("encrypted msg:", encryptedMsgObj)
        current_decrpt, peak_decrpt = tracemalloc.get_traced_memory()
        start_cpu=process.cpu_percent()
        # start_cpu = psutil.cpu_percent()
        start_time = time.perf_counter()
        decryptedMsg = decrypt_RSA(encryptedMsg, privKey)
        siz1=sys.getsizeof(encryptedMsg[0])*8
        end_time = time.perf_counter()
        # end_cpu = psutil.cpu_percent()
        current_decrpt2, peak_decrpt2 = tracemalloc.get_traced_memory()
        end_cpu=process.cpu_percent()
        decry_cpu=end_cpu-start_cpu
        decry_time=end_time-start_time
        decry_mem=current_decrpt2-current_decrpt
        # print("decrypted msg:", decryptedMsg)
        # print(f"Encryption time : {encry_time}")
        # print(f"Decryption time : {decry_time}")
        # print(f"Encryption CPU: {encry_cpu}")
        # print(f"Decryption CPU : {decry_cpu}")
        encryptedMsg1 = encrypt_RSA(modified_msg, pubKey)
        encryptedMsgObj1 = {
            'ciphertext': binascii.hexlify(encryptedMsg1[0]),
            'iv': binascii.hexlify(encryptedMsg1[1]),
            'ciphertextKey': binascii.hexlify(encryptedMsg1[2])
        }
        # print("Length of encrypted msg",len(encryptedMsg[0]))
        # plaintext_entropy = shannon_entropy(msg.decode())
        binary = "".join(f"{x:08b}" for x in encryptedMsg[0])
        with open("rsablow_binary.txt", "w") as f:
            f.write(binary)
        with open("rsablow.txt", "wb") as f:
            f.write((encryptedMsg[0]))
        # convert binary to string

        siz=len(msg)*8
        # siz1=sys.getsizeof(encryptedMsg[0])
        def hamming_distance(a, b):
        # Align the two binary strings by adding leading zeros
            if len(a) > len(b):
                b = '0' * (len(a) - len(b)) + b
            else:
                a = '0' * (len(b) - len(a)) + a

        # Compute the Hamming distance
            distance = 0
            for i in range(len(a)):
                if a[i] != b[i]:
                    distance += 1
            return distance
        binary_data = binascii.unhexlify(encryptedMsgObj['ciphertext'])
        binary_string = ''.join(format(byte, '08b') for byte in binary_data)
        # binary = ''.join(format(ord(char), '08b') for char in string)
        binary_data1 = binascii.unhexlify(encryptedMsgObj1['ciphertext'])
        binary_string1 = ''.join(format(byte, '08b') for byte in binary_data1)
        x=hamming_distance(binary_string,binary_string1)
        # print("Hamming Distance is ")
        # print(x)
        # print("Avalanche Effect: ")
        # print(x/siz)
        # print("Throughput of encryption")
        # print(siz/encry_time)
        # print("Throughput of decryption")
        # print(siz1/decry_time)
        outputs=[file_list[file_count],encry_time,decry_time,encry_cpu,decry_cpu,mem_usage,decry_mem,x,x/siz,siz/encry_time,siz1/decry_time]
        for column, output in enumerate(outputs, start=1):
            sheet.cell(row=row_count, column=column, value=output)
        row_count=row_count+1
        count=count+1
        print("count is",count)
    file_count=file_count+1
    row_count=row_count+10
workbook.save('rsa_blow.xlsx')