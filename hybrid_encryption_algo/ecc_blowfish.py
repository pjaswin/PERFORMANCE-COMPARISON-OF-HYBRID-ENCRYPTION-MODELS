from tinyec import registry
from Crypto.Cipher import Blowfish
import hashlib, secrets, binascii
import time
import psutil
from memory_profiler import profile
import math,sys
import openpyxl
from collections import Counter
import tracemalloc


def encrypt_Blowfish_CBC(msg, secretKey):
    iv = secrets.token_bytes(8)
    cipher = Blowfish.new(secretKey, Blowfish.MODE_CBC, iv)
    padding_length = 8 - (len(msg) % 8)
    padded_msg = msg + bytes([padding_length] * padding_length)
    ciphertext = cipher.encrypt(padded_msg)
    return (ciphertext, iv)

def decrypt_Blowfish_CBC(ciphertext, iv, secretKey):
    cipher = Blowfish.new(secretKey, Blowfish.MODE_CBC, iv)
    padded_msg = cipher.decrypt(ciphertext)
    padding_length = padded_msg[-1]
    return padded_msg[:-padding_length]

def ecc_point_to_256_bit_key(point):
    sha = hashlib.sha256(int.to_bytes(point.x, 32, 'big'))
    sha.update(int.to_bytes(point.y, 32, 'big'))
    return sha.digest()

curve = registry.get_curve('secp256r1')

def encrypt_ECC(msg, pubKey):
    ciphertextPrivKey = secrets.randbelow(curve.field.n)
    sharedECCKey = ciphertextPrivKey * pubKey
    secretKey = ecc_point_to_256_bit_key(sharedECCKey)
    ciphertext, iv = encrypt_Blowfish_CBC(msg, secretKey)
    ciphertextPubKey = ciphertextPrivKey * curve.g
    return (ciphertext, iv, ciphertextPubKey)

def decrypt_ECC(encryptedMsg, privKey):
    (ciphertext, iv, ciphertextPubKey) = encryptedMsg
    sharedECCKey = privKey * ciphertextPubKey
    secretKey = ecc_point_to_256_bit_key(sharedECCKey)
    plaintext = decrypt_Blowfish_CBC(ciphertext, iv, secretKey)
    return plaintext
count=0
workbook = openpyxl.Workbook()
sheet = workbook.active
outputs=["file name","encry_time","decry_time","encry_cpu","decry_cpu","encry_mem","decry_mem","Hamming Distance","Avalache Effect","Through_put encryt","Through_put decrypt"]
file_list=["oneMB.txt","twoMB.txt","fiveMB.txt","tenMB.txt","twentyMB.txt"]
file_count=0
row_count=2
for column, output in enumerate(outputs, start=1):
    sheet.cell(row=1, column=column, value=output)
while file_count<5:#changed
    count=0
    while count<30:#changed
        with open(file_list[file_count], 'r') as file:
            msg = file.read()
        modified_msg = msg[:2] + 'XY' + msg[4:]
        msg=msg.encode()
        modified_msg=modified_msg.encode()
        process = psutil.Process()
        # msg = b'Text to be encrypted by ECC public key and ' \
        #       b'decrypted by its corresponding ECC private key'
        # print("original msg:", msg)
        tracemalloc.start()
        current_encrpt, peak_encrpt = tracemalloc.get_traced_memory()
        start_cpu = process.cpu_percent()
        start_time = time.perf_counter()
        privKey = secrets.randbelow(curve.field.n)
        # privKey=40430284738272264068566731406742247973237551945965824218361646746498880385806
        # privKey=40430284738272264068566731406742247973237551945965824218361646746498880385807
        pubKey = privKey * curve.g

        encryptedMsg = encrypt_ECC(msg, pubKey)
        end_time = time.perf_counter()
        encry_time=end_time-start_time
        end_cpu = process.cpu_percent()
        encry_cpu=end_cpu-start_cpu
        current_encrpt2, peak_encrpt2 = tracemalloc.get_traced_memory()
        mem_usage=current_encrpt2-current_encrpt
        encryptedMsgObj = {
            'ciphertext': binascii.hexlify(encryptedMsg[0]),
            'iv': binascii.hexlify(encryptedMsg[1]),
            'ciphertextPubKey': hex(encryptedMsg[2].x) + hex(encryptedMsg[2].y % 2)[2:]
        }
        # print("encrypted msg:", encryptedMsgObj)
        tracemalloc.start()
        current_decrpt, peak_decrpt = tracemalloc.get_traced_memory()
        start_cpu = process.cpu_percent()
        start_time = time.perf_counter()
        decryptedMsg = decrypt_ECC(encryptedMsg, privKey)
        end_time = time.perf_counter()
        end_cpu = process.cpu_percent()
        current_decrpt2, peak_decrpt2 = tracemalloc.get_traced_memory()
        decry_cpu=end_cpu-start_cpu
        decry_time=end_time-start_time
        decry_mem=current_decrpt2-current_decrpt
        # print("decrypted msg:", decryptedMsg)
        # print(f"Encryption time : {encry_time}")
        # print(f"Decryption time : {decry_time}")
        # print(f"Encryption CPU: {encry_cpu}")
        # print(f"Decryption CPU : {decry_cpu}")
        encryptedMsg1 = encrypt_ECC(modified_msg, pubKey)
        encryptedMsgObj1 = {
            'ciphertext': binascii.hexlify(encryptedMsg1[0]),
            'iv': binascii.hexlify(encryptedMsg1[1]),
            'ciphertextPubKey': hex(encryptedMsg1[2].x) + hex(encryptedMsg1[2].y % 2)[2:]
        }
        siz=len(msg)*8
        siz1=sys.getsizeof(encryptedMsg[0])*8
        binary = "".join(f"{x:08b}" for x in encryptedMsg[0])
        # with open("sec1.txt", "wb") as f:
        #     f.write(encryptedMsg[0])
        with open("eccblow.txt", "wb") as f:
            f.write(encryptedMsg[0])
        with open("eccblow_binary.txt", "w") as f:
            f.write(binary)
        # convert binary to string

        def hamming_distance(a, b):
        # Align the two binary strings by adding leading zeros
            if len(a) > len(b):
                b = '0' * (len(a) - len(b)) + b
            else:
                a = '0' * (len(b) - len(a)) + a

        # Compute the Hamming distance
            distance = 0
            for i in range(len(a)):
                if a[i] != b[i]:
                    distance += 1
            return distance
        binary_data = binascii.unhexlify(encryptedMsgObj['ciphertext'])
        binary_string = ''.join(format(byte, '08b') for byte in binary_data)
        binary_data1 = binascii.unhexlify(encryptedMsgObj1['ciphertext'])
        binary_string1 = ''.join(format(byte, '08b') for byte in binary_data1)
        x=hamming_distance(binary_string,binary_string1)
        outputs=[file_list[file_count],encry_time,decry_time,encry_cpu,decry_cpu,mem_usage,decry_mem,x,x/siz,siz/encry_time,siz1/decry_time]
        for column, output in enumerate(outputs, start=1):
            sheet.cell(row=row_count, column=column, value=output)
        row_count=row_count+1
        count=count+1
        print("count is",count)
    file_count=file_count+1
    row_count=row_count+10
        # print("Hamming Distance is ")
        # print(x)
        # print("Avalanche Effect: ")
        # print(x/siz)
        # print("Throughput of encryption")
        # print(siz/encry_time)
        # print("Throughput of decryption")
        # print(siz1/decry_time)
workbook.save('ecc_blow.xlsx')